from flask import Flask, request, jsonify, render_template_string, Response
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from pathlib import Path

app = Flask(__name__)
@app.route('/api/logs', methods=['GET'])
def get_logs():
    ensure_db()
    wb = load_workbook(DB_PATH)
    ws = wb['logs']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify([])
    header, data = rows[0], rows[1:]
    # Laatste 100 logs, nieuwste eerst
    data = data[-100:][::-1]
    logs = [dict(zip(header, row)) for row in data]
    return jsonify(logs)
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
from pathlib import Path

app = Flask(__name__)

BASE_DIR = Path(__file__).resolve().parent.parent
DB_PATH = Path(os.getenv('LOG_DB_PATH', str(BASE_DIR / 'Database' / 'logs.xlsx')))
LOG_COLUMNS = ['datetime', 'what', 'module', 'where']

# Ensure database directory exists
def ensure_db():
    os.makedirs(DB_PATH.parent, exist_ok=True)
    if not DB_PATH.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = 'logs'
        ws.append(LOG_COLUMNS)
        wb.save(DB_PATH)

@app.route('/api/log', methods=['POST'])
def log():
    ensure_db()
    data = request.json
    row = [
        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        data.get('what', ''),
        data.get('module', ''),
        data.get('where', '')
    ]
    wb = load_workbook(DB_PATH)
    ws = wb['logs']
    ws.append(row)
    wb.save(DB_PATH)
    return jsonify({'status': 'ok'})

@app.route('/search', methods=['GET'])
def search():
    ensure_db()
    query = request.args.get('q', '').lower()
    wb = load_workbook(DB_PATH)
    ws = wb['logs']
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        rows = [tuple(LOG_COLUMNS)]
    header, data = rows[0], rows[1:]
    if query:
        data = [r for r in data if any(query in str(cell).lower() for cell in r)]
    html = '''
    <h2>Log Search</h2>
    <form method="get">
        <input name="q" value="{{query}}" placeholder="Search logs">
        <button type="submit">Search</button>
    </form>
    <table border=1 cellpadding=4>
        <tr>{% for col in header %}<th>{{col}}</th>{% endfor %}</tr>
        {% for row in data %}<tr>{% for cell in row %}<td>{{cell}}</td>{% endfor %}</tr>{% endfor %}
    </table>
    '''
    return render_template_string(html, header=header, data=data, query=query)

if __name__ == '__main__':
    ensure_db()
    app.run(host='0.0.0.0', port=5001)
