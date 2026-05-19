from __future__ import annotations
import os
import queue
import socket
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
import webbrowser
from dataclasses import dataclass
from pathlib import Path
from tkinter import END, BOTH, LEFT, RIGHT, VERTICAL, Y, Tk, ttk, Text, messagebox, BooleanVar

BASE_DIR = Path(__file__).resolve().parents[1]
PYTHON_EXE = sys.executable

@dataclass
class Service:
    name: str
    command: list[str]
    cwd: Path
    is_docker: bool = False
    env: dict[str, str] | None = None
    process: subprocess.Popen | None = None
    startup_state: str = "offline"
    run_mode: str | None = None

    def is_running(self) -> bool:
        if self.is_docker:
            compose_file = self.cwd / "docker-compose.yml"
            if compose_file.exists():
                try:
                    result = subprocess.run(
                        ["docker", "compose", "ps", "--status", "running", "logging-service"],
                        cwd=self.cwd,
                        capture_output=True,
                        text=True,
                        encoding='utf-8',
                        errors='replace',
                        timeout=5,
                    )
                    return "logging-service" in result.stdout
                except:
                    return False
            # Fallback: old Kubernetes check
            try:
                result = subprocess.run(
                    ["kubectl", "get", "pods", "-n", "logging-center", "-l", "app=logging-service", "--field-selector=status.phase=Running", "-o", "jsonpath={.items[*].metadata.name}"],
                    capture_output=True,
                    text=True,
                    encoding='utf-8',
                    errors='replace',
                    timeout=5
                )
                return len(result.stdout.strip()) > 0
            except:
                return False
        return self.process is not None and self.process.poll() is None


class StartupApp:
    def __init__(self, root: Tk) -> None:
        self.root = root
        self.root.title("Startup Beheer")
        self.root.geometry("900x520")

        self.log_queue: queue.Queue[str] = queue.Queue()
        self.services: list[Service] = self._load_services()
        self.api_enabled: bool = True

        self._build_ui()
        self._render_services()
        self._refresh_status()
        self._drain_log_queue()

    def _load_services(self) -> list[Service]:
        services: list[Service] = []

        # Server Website
        server_app = BASE_DIR / "Server" / "app.py"
        if server_app.exists():
            services.append(
                Service(
                    name="Server Website",
                    command=[PYTHON_EXE, str(server_app)],
                    cwd=server_app.parent,
                )
            )

        # Logging Service (mode will be chosen at start: Docker or Local Python)
        logging_service = BASE_DIR / "Logging Centrum" / "logging_service.py"
        logging_dir = BASE_DIR / "Logging Centrum"
        docker_compose = logging_dir / "docker-compose.yml"
        if logging_service.exists() or docker_compose.exists():
            services.append(
                Service(
                    name="Logging Service",
                    command=[],
                    cwd=logging_dir,
                )
            )

        return services

    def _build_ui(self) -> None:
        container = ttk.Frame(self.root, padding=12)
        container.pack(fill=BOTH, expand=True)

        header_frame = ttk.Frame(container)
        header_frame.pack(fill=BOTH)
        ttk.Label(header_frame, text="Startup Beheer", font=("Segoe UI", 12, "bold")).pack(
            side=LEFT
        )

        top_frame = ttk.Frame(container)
        top_frame.pack(fill=BOTH, expand=True)

        self.tree = ttk.Treeview(
            top_frame,
            columns=("indicator", "status", "path"),
            show="headings",
            height=10,
        )
        self.tree.heading("indicator", text="●")
        self.tree.heading("status", text="Status")
        self.tree.heading("path", text="Entry")
        self.tree.column("indicator", width=40, anchor="center")
        self.tree.column("status", width=120, anchor="center")
        self.tree.column("path", width=480)
        self.tree.pack(side=LEFT, fill=BOTH, expand=True)
        self.tree.tag_configure("online", foreground="#22c55e")
        self.tree.tag_configure("starting", foreground="#eab308")
        self.tree.tag_configure("offline", foreground="#ef4444")

        scrollbar = ttk.Scrollbar(top_frame, orient=VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        scrollbar.pack(side=RIGHT, fill=Y)

        button_frame = ttk.Frame(container)
        button_frame.pack(fill=BOTH, pady=(12, 6))

        ttk.Button(button_frame, text="Start", command=self.start_selected).pack(
            side=LEFT, padx=6
        )
        ttk.Button(button_frame, text="Stop", command=self.stop_selected).pack(
            side=LEFT, padx=6
        )
        ttk.Button(button_frame, text="Restart", command=self.restart_selected).pack(
            side=LEFT, padx=6
        )
        ttk.Separator(button_frame, orient="vertical").pack(side=LEFT, fill=Y, padx=8)
        ttk.Button(button_frame, text="Start All", command=self.start_all).pack(
            side=LEFT, padx=6
        )
        ttk.Button(button_frame, text="Stop All", command=self.stop_all).pack(
            side=LEFT, padx=6
        )
        ttk.Button(button_frame, text="Restart All", command=self.restart_all).pack(
            side=LEFT, padx=6
        )
        ttk.Separator(button_frame, orient="vertical").pack(side=LEFT, fill=Y, padx=8)
        ttk.Button(button_frame, text="Log Search", command=self.open_log_search).pack(
            side=LEFT, padx=6
        )
        ttk.Separator(button_frame, orient="vertical").pack(side=LEFT, fill=Y, padx=8)

        self.api_enabled_var = BooleanVar(value=True)
        api_check = ttk.Checkbutton(
            button_frame,
            text="API System Enabled",
            variable=self.api_enabled_var,
            command=self._on_api_toggle,
        )
        api_check.pack(side=LEFT, padx=6)

        log_label = ttk.Label(container, text="Logs")
        log_label.pack(anchor="w", pady=(6, 0))

        self.log_text = Text(container, height=10, wrap="word")
        self.log_text.pack(fill=BOTH, expand=True)

    def _on_api_toggle(self) -> None:
        self.api_enabled = self.api_enabled_var.get()
        status = "ingeschakeld" if self.api_enabled else "uitgeschakeld"
        self._log(f"API System {status}.")
        threading.Thread(target=self._send_api_toggle, daemon=True).start()

    def _send_api_toggle(self) -> None:
        try:
            import json
            payload = json.dumps({"enabled": self.api_enabled}).encode("utf-8")
            request_obj = urllib.request.Request(
                "http://127.0.0.1:7000/api/system/toggle",
                data=payload,
                headers={"Content-Type": "application/json"},
                method="POST",
            )
            with urllib.request.urlopen(request_obj, timeout=2.0):
                self._log(f"✓ API System toggle synchroniseerd met server.")
        except Exception as e:
            self._log(f"✗ Kon API toggle niet synchroniseren: {str(e)}")

    def open_log_search(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        try:
            from openpyxl import load_workbook
        except ImportError:
            messagebox.showerror("Dependency missing", "Install openpyxl to use log search.")
            return
        log_path = (BASE_DIR / "Database" / "logs.xlsx")
        if not log_path.exists():
            messagebox.showinfo("Geen logs", "Geen logbestand gevonden.")
            return
        search_win = tk.Toplevel(self.root)
        search_win.title("Log Search")
        search_win.geometry("900x500")

        search_frame = ttk.Frame(search_win, padding=10)
        search_frame.pack(fill=tk.BOTH, expand=True)

        search_var = tk.StringVar()
        ttk.Label(search_frame, text="Zoekterm:").pack(side=tk.LEFT)
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=40)
        search_entry.pack(side=tk.LEFT, padx=6)
        tree = ttk.Treeview(search_win, show="headings")
        tree.pack(fill=tk.BOTH, expand=True, pady=(10,0))

        def load_logs():
            wb = load_workbook(log_path)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messagebox.showinfo("Geen logs", "Geen logbestand gevonden of leeg.")
                return
            header, data = rows[0], rows[1:]
            tree.delete(*tree.get_children())
            tree['columns'] = header
            for col in header:
                tree.heading(col, text=col)
                tree.column(col, width=120, anchor="center")
            q = search_var.get().lower()
            for row in data:
                if not q or any(q in str(cell).lower() for cell in row):
                    tree.insert('', 'end', values=row)

        ttk.Button(search_frame, text="Zoek", command=load_logs).pack(side=tk.LEFT, padx=6)
        load_logs()

    def _render_services(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for idx, service in enumerate(self.services):
            if service.is_running():
                service.startup_state = "online"
                status = "Online"
                tag = "online"
            elif service.startup_state == "starting":
                status = "Starting"
                tag = "starting"
            else:
                service.startup_state = "offline"
                status = "Offline"
                tag = "offline"
            # Show service name instead of command
            entry = service.name
            self.tree.insert("", END, iid=str(idx), values=("●", status, entry), tags=(tag,))

    def _refresh_status(self) -> None:
        for idx, service in enumerate(self.services):
            if not self.tree.exists(str(idx)):
                continue
            if service.is_running():
                service.startup_state = "online"
                status = "Online"
                tag = "online"
            elif service.startup_state == "starting":
                status = "Starting"
                tag = "starting"
            else:
                service.startup_state = "offline"
                status = "Offline"
                tag = "offline"
            self.tree.set(str(idx), "indicator", "●")
            self.tree.set(str(idx), "status", status)
            self.tree.item(str(idx), tags=(tag,))
        self.root.after(1000, self._refresh_status)

    def _log(self, message: str) -> None:
        timestamp = time.strftime("%H:%M:%S")
        self.log_queue.put(f"[{timestamp}] {message}\n")

    def _drain_log_queue(self) -> None:
        while not self.log_queue.empty():
            line = self.log_queue.get_nowait()
            self.log_text.insert(END, line)
            self.log_text.see(END)
        self.root.after(200, self._drain_log_queue)

    def _start_service(self, service: Service) -> None:
        if service.is_running():
            self._log(f"{service.name} is al gestart.")
            return
        if service.name == "Logging Service":
            selected_mode = self._choose_logging_mode()
            if selected_mode is None:
                self._log("Start Logging Service geannuleerd.")
                return
            logging_py = BASE_DIR / "Logging Centrum" / "logging_service.py"
            if selected_mode == "docker":
                service.is_docker = True
                service.run_mode = "docker"
                service.command = ["docker", "compose", "up", "-d", "--build"]
                service.cwd = BASE_DIR / "Logging Centrum"
            else:
                if not logging_py.exists():
                    self._log("logging_service.py niet gevonden voor Local Python mode.")
                    return
                service.is_docker = False
                service.run_mode = "local"
                service.command = [PYTHON_EXE, str(logging_py)]
                service.cwd = logging_py.parent
        service.startup_state = "starting"
        try:
            if service.is_docker:
                self._log(f"{service.name} wordt gestart...")
                self._log("⏳ Docker compose wordt uitgevoerd...")
                service.process = subprocess.Popen(
                    service.command,
                    cwd=service.cwd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8',
                    errors='replace',
                )
                # Wait for kubectl to finish
                threading.Thread(
                    target=self._wait_docker_start,
                    args=(service,),
                    daemon=True,
                ).start()
            else:
                # Regular Python service
                run_env = os.environ.copy()
                if service.env:
                    run_env.update(service.env)
                service.process = subprocess.Popen(
                    service.command,
                    cwd=service.cwd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding='utf-8',
                    errors='replace',
                    env=run_env,
                )
                if service.name != "Server Website":
                    self._log(f"{service.name} gestart.")
                threading.Thread(
                    target=self._stream_output,
                    args=(service,),
                    daemon=True,
                ).start()
        except Exception as exc:
            service.startup_state = "offline"
            self._log(f"Fout bij starten van {service.name}: {exc}")

    def _wait_docker_start(self, service: Service) -> None:
        """Wait for docker-compose command to complete and stream output"""
        if service.process is None or service.process.stdout is None:
            return
        # Stream output during build/start
        for line in service.process.stdout:
            cleaned = line.strip()
            if cleaned:
                self._log(f"{service.name}: {cleaned}")
        # Check return code
        service.process.wait()
        if service.process.returncode == 0:
            service.startup_state = "online"
            self._log(f"{service.name} succesvol gestart!")
            if service.name == "Logging Service":
                server_ip = self._get_server_ip()
                self._log(f"{service.name}: bereikbaar op http://{server_ip}:5001")
                ok, details = self._check_website_api_status(server_ip)
                state_label = "Succes" if ok else "Failed"
                self._log(f"API check http://{server_ip}:7000/api/permissions => {state_label} ({details})")
        else:
            service.startup_state = "offline"
            self._log(f"{service.name} fout bij starten (exit code: {service.process.returncode})")
        service.process = None  # Clear process after docker-compose finishes

    def _stream_output(self, service: Service) -> None:
        if service.process is None or service.process.stdout is None:
            return
        for line in service.process.stdout:
            cleaned = line.strip()
            if not cleaned:
                continue
            if service.name == "Logging Service" and "Running on http://" in cleaned:
                try:
                    server_ip = self._get_server_ip()
                    normalized = f" * Running on http://{server_ip}:5001"
                    self._log(f"{service.name}: {normalized}")
                except Exception:
                    self._log(f"{service.name}: {cleaned}")
                continue
            if service.name == "Server Website":
                if "Running on http" in cleaned:
                    try:
                        url = cleaned.split("Running on ")[-1].strip()
                        if url.startswith("http://127.0.0.1") or url.startswith("http://localhost"):
                            continue
                        self._log(f"{service.name}: {cleaned}")
                        threading.Thread(target=self._open_chrome, args=(url,), daemon=True).start()
                        self._log(f"Chrome geopend: {url}")
                    except Exception as e:
                        self._log(f"Fout bij openen Chrome: {e}")
                elif "Serving Flask app" in cleaned:
                    self._log(f"{service.name}: gestart")
                continue
            self._log(f"{service.name}: {cleaned}")

    def _get_server_ip(self) -> str:
        """Best-effort detection of the host machine LAN IP."""
        try:
            sock = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            sock.connect(("8.8.8.8", 80))
            ip = sock.getsockname()[0]
            sock.close()
            if ip and not ip.startswith("127."):
                return ip
        except Exception:
            pass
        return "127.0.0.1"

    def _check_website_api_status(self, server_ip: str) -> tuple[bool, str]:
        """Probe website API on server IP and return success/failure with details."""
        url = f"http://{server_ip}:7000/api/permissions"
        try:
            req = urllib.request.Request(url, method="GET")
            with urllib.request.urlopen(req, timeout=3) as response:
                code = response.getcode()
                return (200 <= code < 400), f"HTTP {code}"
        except urllib.error.HTTPError as exc:
            # 401/403 mean API is reachable but access is restricted (still success for connectivity)
            if exc.code in {401, 403}:
                return True, f"HTTP {exc.code}"
            return False, f"HTTP {exc.code}"
        except Exception as exc:
            return False, str(exc)

    def _open_chrome(self, url: str) -> None:
        """Open URL in Google Chrome"""
        try:
            import winreg
            # Find Chrome installation path
            chrome_path = None
            try:
                reg_path = r"SOFTWARE\Microsoft\Windows\CurrentVersion\App Paths\chrome.exe"
                with winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE, reg_path) as key:
                    chrome_path = winreg.QueryValue(key, None)
            except:
                pass
            
            if chrome_path and Path(chrome_path).exists():
                subprocess.Popen([chrome_path, url])
            else:
                # Fallback to webbrowser if Chrome not found
                webbrowser.open(url)
        except Exception as e:
            self._log(f"Chrome open fout: {e}")
            webbrowser.open(url)

    def _stop_service(self, service: Service) -> None:
        if not service.is_running():
            self._log(f"{service.name} draait niet.")
            return
        
        if service.is_docker:
            try:
                self._log(f"{service.name} wordt gestopt...")
                compose_file = service.cwd / "docker-compose.yml"
                if compose_file.exists():
                    result = subprocess.run(
                        ["docker", "compose", "down"],
                        cwd=service.cwd,
                        capture_output=True,
                        text=True,
                        encoding='utf-8',
                        errors='replace',
                        timeout=30,
                    )
                else:
                    k8s_manifest = BASE_DIR / "Logging Centrum" / "k8s-local.yaml"
                    result = subprocess.run(
                        ["kubectl", "delete", "-f", str(k8s_manifest)],
                        cwd=service.cwd,
                        capture_output=True,
                        text=True,
                        encoding='utf-8',
                        errors='replace',
                        timeout=30,
                    )
                if result.returncode == 0:
                    service.startup_state = "offline"
                    self._log(f"{service.name} gestopt.")
                else:
                    self._log(f"{service.name} fout bij stoppen: {result.stderr}")
            except Exception as exc:
                self._log(f"Fout bij stoppen van {service.name}: {exc}")
        else:
            # Stop regular Python service
            assert service.process is not None
            service.process.terminate()
            try:
                service.process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                service.process.kill()
            service.startup_state = "offline"
            self._log(f"{service.name} gestopt.")

    def _restart_service(self, service: Service) -> None:
        self._stop_service(service)
        self._start_service(service)

    def _choose_logging_mode(self) -> str | None:
        choice = messagebox.askyesnocancel(
            "Logging Service mode",
            "Kies mode voor Logging Service:\n\nJa = Docker run\nNee = Local Python run\nAnnuleren = stop",
        )
        if choice is None:
            return None
        if choice:
            return "docker"
        return "local"


    def _get_selected_services(self) -> list[Service]:
        selected = self.tree.selection()
        if not selected:
            messagebox.showinfo("Geen selectie", "Selecteer een service.")
            return []
        return [self.services[int(idx)] for idx in selected]

    def start_selected(self) -> None:
        for service in self._get_selected_services():
            self._start_service(service)

    def stop_selected(self) -> None:
        for service in self._get_selected_services():
            self._stop_service(service)

    def restart_selected(self) -> None:
        for service in self._get_selected_services():
            self._restart_service(service)

    def start_all(self) -> None:
        for service in self.services:
            self._start_service(service)

    def stop_all(self) -> None:
        for service in self.services:
            self._stop_service(service)

    def restart_all(self) -> None:
        for service in self.services:
            self._restart_service(service)


if __name__ == "__main__":
    root = Tk()
    app = StartupApp(root)
    if not app.services:
        messagebox.showwarning(
            "Geen services",
            "Geen startbare services gevonden in de projectmap.",
        )
    root.mainloop()
